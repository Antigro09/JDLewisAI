"""Excel export via openpyxl (real implementation).

Workbook layout:
  Summary   — totals by item type/unit
  Takeoff   — one row per quantity with formula, confidence, review state
  Audit     — evidence chain per quantity (geometry/OCR/scale ids)
"""

from __future__ import annotations

from datetime import UTC, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.adapters.base import ExportAdapter
from app.export.audit import audit_summary

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_REVIEW_FILL = PatternFill("solid", fgColor="FEF3C7")


def _header(ws, cols: list[str], widths: list[int]):
    ws.append(cols)
    for i, w in enumerate(widths, start=1):
        cell = ws.cell(row=1, column=i)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        ws.column_dimensions[cell.column_letter].width = w
    ws.freeze_panes = "A2"


class ExcelExportAdapter(ExportAdapter):
    name = "excel-export"

    def export(self, project_id: str, payload: dict[str, Any], out_path: str) -> str:
        quantities: list[dict] = payload["quantities"]
        sheets_by_id = {s["id"]: s for s in payload.get("sheets", [])}

        wb = Workbook()

        # --- Summary -------------------------------------------------------
        ws = wb.active
        ws.title = "Summary"
        _header(ws, ["Item Type", "CSI", "Unit", "Total Quantity", "Items", "Needs Review"],
                [24, 12, 8, 16, 8, 14])
        totals: dict[tuple[str, str, str], dict] = {}
        for q in quantities:
            key = (q["item_type"], q.get("csi_code") or "", q["unit"])
            t = totals.setdefault(key, {"qty": 0.0, "n": 0, "review": 0})
            t["qty"] += q["quantity"]
            t["n"] += 1
            t["review"] += 1 if q["needs_review"] else 0
        for (item_type, csi, unit), t in sorted(totals.items()):
            ws.append([item_type, csi, unit, round(t["qty"], 2), t["n"], t["review"]])
        ws.append([])
        ws.append([
            f"Generated {datetime.now(UTC):%Y-%m-%d %H:%M UTC} — machine-assisted "
            "takeoff; review-flagged rows are highlighted on the Takeoff tab."
        ])

        # --- Takeoff ---------------------------------------------------------
        ws = wb.create_sheet("Takeoff")
        cols = ["Sheet", "Page", "Item Type", "Description", "Quantity", "Unit", "CSI",
                "Formula", "Scale Conf", "Meas Conf", "Model Conf", "Final Conf",
                "Needs Review", "Review Reasons", "Status", "Audit Notes", "Item ID"]
        _header(ws, cols, [10, 6, 18, 30, 12, 6, 10, 60, 10, 10, 10, 10, 12, 30, 10, 42, 34])
        for q in quantities:
            sheet = sheets_by_id.get(q["sheet_id"], {})
            row = [
                sheet.get("sheet_number") or sheet.get("id", "")[:8],
                q["page_number"], q["item_type"], q["description"],
                q["quantity"], q["unit"], q.get("csi_code") or "",
                q["formula"],
                round(q.get("scale_confidence", 0), 2),
                round(q.get("measurement_confidence", 0), 2),
                round(q.get("model_confidence", 0), 2),
                round(q.get("final_confidence", 0), 2),
                "YES" if q["needs_review"] else "",
                ", ".join(q.get("review_reason", [])),
                q.get("review_status", "pending"),
                audit_summary(q),
                q["id"],
            ]
            ws.append(row)
            if q["needs_review"]:
                for cell in ws[ws.max_row]:
                    cell.fill = _REVIEW_FILL

        # --- Audit -------------------------------------------------------------
        ws = wb.create_sheet("Audit")
        _header(ws, ["Item ID", "Scale ID", "Geometry IDs", "OCR Span IDs", "Attributes"],
                [34, 34, 50, 50, 60])
        for q in quantities:
            ws.append([
                q["id"], q.get("scale_id") or "",
                ", ".join(q.get("source_geometry_ids", [])),
                ", ".join(q.get("source_ocr_span_ids", [])),
                str(q.get("attributes", {})),
            ])

        wb.save(out_path)
        return out_path
