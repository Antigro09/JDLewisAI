import csv

from openpyxl import load_workbook

from app.export.audit import audit_summary
from app.export.csv_export import CSVExportAdapter
from app.export.excel import ExcelExportAdapter


def _payload():
    return {
        "sheets": [{"id": "s1", "sheet_number": "A1.00"}],
        "quantities": [
            {
                "id": "door",
                "sheet_id": "s1",
                "page_number": 5,
                "item_type": "door",
                "description": "Doors (symbol count)",
                "quantity": 22,
                "unit": "EA",
                "csi_code": "08 11 00",
                "formula": "EA = count of accepted door symbols = 22",
                "scale_confidence": 1.0,
                "measurement_confidence": 1.0,
                "model_confidence": 0.82,
                "final_confidence": 0.91,
                "needs_review": False,
                "review_reason": [],
                "review_status": "pending",
                "source_geometry_ids": ["d1", "d2"],
                "source_ocr_span_ids": ["o1"],
                "attributes": {
                    "symbol_count": 22,
                    "unique_mark_count": 17,
                    "mark_counts": {"100A": 2, "101": 1},
                },
            },
            {
                "id": "wall",
                "sheet_id": "s1",
                "page_number": 5,
                "item_type": "wall",
                "description": "Wall S2-0-6",
                "quantity": 94.68,
                "unit": "LF",
                "csi_code": "09 22 16",
                "formula": "LF = sum of 14 wall segments",
                "scale_confidence": 1.0,
                "measurement_confidence": 1.0,
                "model_confidence": 0.75,
                "final_confidence": 0.9,
                "needs_review": False,
                "review_reason": [],
                "review_status": "pending",
                "source_geometry_ids": ["g1", "g2"],
                "source_ocr_span_ids": [],
                "attributes": {
                    "wall_code": "S2-0-6",
                    "segment_count": 14,
                    "segment_lengths_lf": [14.73, 22.79],
                    "unit_size_in": 6,
                },
            },
        ],
    }


def test_csv_export_includes_flat_audit_notes_and_attributes(tmp_path):
    out = tmp_path / "takeoff.csv"

    CSVExportAdapter().export("p1", _payload(), str(out))

    rows = list(csv.DictReader(out.open()))
    assert rows[0]["audit_notes"] == (
        "22 accepted symbols; 17 unique scheduled marks; duplicate marks: 100A x2"
    )
    assert rows[0]["source_geometry_ids"] == "d1,d2"
    assert '"symbol_count": 22' in rows[0]["attributes"]
    assert rows[1]["audit_notes"].startswith("14 wall segments; segment LF: 14.73, 22.79")


def test_excel_takeoff_sheet_includes_audit_notes(tmp_path):
    out = tmp_path / "takeoff.xlsx"

    ExcelExportAdapter().export("p1", _payload(), str(out))

    wb = load_workbook(out, read_only=True)
    ws = wb["Takeoff"]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    row = [cell.value for cell in next(ws.iter_rows(min_row=2, max_row=2))]
    audit_idx = header.index("Audit Notes")

    assert row[audit_idx] == (
        "22 accepted symbols; 17 unique scheduled marks; duplicate marks: 100A x2"
    )


def test_schedule_backed_door_audit_explains_existing_exclusion():
    summary = audit_summary({
        "item_type": "door",
        "attributes": {
            "count_basis": "scheduled_plan_marks",
            "opening_count": 19,
            "schedule_row_count": 20,
            "existing_schedule_marks_excluded": ["100A"],
        },
    })

    assert summary == (
        "19 scheduled openings; 20 door schedule rows; existing/ETR excluded: 100A"
    )
